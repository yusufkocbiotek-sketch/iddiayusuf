import json
import os
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


JSON_DOSYA = r"C:\Users\YUSUF\Desktop\iddiayusuf-main\public\data\mac.json"
EXCEL_DOSYA = r"C:\Users\YUSUF\Desktop\iddiayusuf-main\public\data\maclar_sablon.xlsx"

# Bu dosyaya ekran görüntüsündeki listeyi yapıştırabilirsin.
# Her satır şöyle olmalı:
# Maç Sonucu_1    A1
# Maç Sonucu_0    A2
# Maç Sonucu_2    A3
SABLON_TXT = r"C:\Users\YUSUF\Desktop\iddiayusuf-main\public\data\oran_sablonu.txt"


# Eğer oran_sablonu.txt yoksa aşağıdaki liste kullanılır.
# Buraya kendi listenin tamamını ekleyebilirsin.
VARSAYILAN_SABLON = """
Maç Sonucu_1	A1
Maç Sonucu_0	A2
Maç Sonucu_2	A3
Handikaplı Maç Sonucu 1:0_1	B1
Handikaplı Maç Sonucu 1:0_0	B2
Handikaplı Maç Sonucu 1:0_2	B3
Handikaplı Maç Sonucu 0:1_1	B9
Handikaplı Maç Sonucu 0:1_0	B10
Handikaplı Maç Sonucu 0:1_2	B11
Maç Sonucu ve Alt/Üst 1.5_1 ve Alt	C1
Maç Sonucu ve Alt/Üst 1.5_0 ve Alt	C2
Maç Sonucu ve Alt/Üst 1.5_2 ve Alt	C3
Maç Sonucu ve Alt/Üst 1.5_1 ve Üst	C4
Maç Sonucu ve Alt/Üst 1.5_0 ve Üst	C5
Maç Sonucu ve Alt/Üst 1.5_2 ve Üst	C6
Maç Sonucu ve Alt/Üst 2.5_1 ve Alt	C7
Maç Sonucu ve Alt/Üst 2.5_0 ve Alt	C8
Maç Sonucu ve Alt/Üst 2.5_2 ve Alt	C9
Maç Sonucu ve Alt/Üst 2.5_1 ve Üst	C10
Maç Sonucu ve Alt/Üst 2.5_0 ve Üst	C11
Maç Sonucu ve Alt/Üst 2.5_2 ve Üst	C12
Maç Sonucu ve Alt/Üst 3.5_1 ve Alt	C13
Maç Sonucu ve Alt/Üst 3.5_0 ve Alt	C14
Maç Sonucu ve Alt/Üst 3.5_2 ve Alt	C15
Maç Sonucu ve Alt/Üst 3.5_1 ve Üst	C16
Maç Sonucu ve Alt/Üst 3.5_0 ve Üst	C17
Maç Sonucu ve Alt/Üst 3.5_2 ve Üst	C18
Ev Sahibi Her İki Yarıyı Kazanır_Evet	D1
Deplasman Her İki Yarıyı Kazanır_Evet	D2
İlk Yarı / Maç Sonucu_1/1	E1
İlk Yarı / Maç Sonucu_1/0	E2
İlk Yarı / Maç Sonucu_1/2	E3
İlk Yarı / Maç Sonucu_0/1	E4
İlk Yarı / Maç Sonucu_0/0	E5
İlk Yarı / Maç Sonucu_0/2	E6
İlk Yarı / Maç Sonucu_2/1	E7
İlk Yarı / Maç Sonucu_2/0	E8
İlk Yarı / Maç Sonucu_2/2	E9
İlk Yarı Sonucu_1	F1
İlk Yarı Sonucu_0	F2
İlk Yarı Sonucu_2	F3
Karşılıklı Gol_Var	G1
Karşılıklı Gol_Yok	G2
Alt/Üst 1.5_Alt	H1
Alt/Üst 1.5_Üst	H2
Alt/Üst 2.5_Alt	H3
Alt/Üst 2.5_Üst	H4
Alt/Üst 3.5_Alt	H5
Alt/Üst 3.5_Üst	H6
İkinci Yarı Sonucu_1	J1
İkinci Yarı Sonucu_0	J2
İkinci Yarı Sonucu_2	J3
İlk Yarı Karşılıklı Gol_Var	M1
İlk Yarı Karşılıklı Gol_Yok	M3
Maç Sonucu ve Karşılıklı Gol_1 ve Var	P1
Maç Sonucu ve Karşılıklı Gol_1 ve Yok	P3
Maç Sonucu ve Karşılıklı Gol_0 ve Var	P4
Maç Sonucu ve Karşılıklı Gol_0 ve Yok	P5
Maç Sonucu ve Karşılıklı Gol_2 ve Var	P6
Maç Sonucu ve Karşılıklı Gol_2 ve Yok	P7
Toplam Gol_0-1 gol	R1
Toplam Gol_2-3 gol	R2
Toplam Gol_4-5 gol	R3
Toplam Gol_6+ gol	R4
İlk Golü Hangi Takım Atar_1	S1
İlk Golü Hangi Takım Atar_Gol Olmaz	S2
İlk Golü Hangi Takım Atar_2	S3
Alt/Üst 0.5_Alt	AA1
Alt/Üst 1.5_Alt	AA2
Alt/Üst 1.5_Üst	AA3
Alt/Üst 2.5_Alt	AA4
Alt/Üst 2.5_Üst	AA5
Alt/Üst 3.5_Alt	AA6
Alt/Üst 3.5_Üst	AA7
Maç Skoru_1:0	FF1
Maç Skoru_2:0	FF2
Maç Skoru_2:1	FF3
Maç Skoru_3:0	FF4
Maç Skoru_3:1	FF5
Maç Skoru_3:2	FF6
Maç Skoru_4:0	FF7
Maç Skoru_4:1	FF8
Maç Skoru_4:2	FF9
Maç Skoru_5:1	FF10
Maç Skoru_0:0	FF11
Maç Skoru_1:1	FF12
Maç Skoru_2:2	FF13
Maç Skoru_3:3	FF14
Maç Skoru_0:1	FF15
Maç Skoru_0:2	FF16
Maç Skoru_1:2	FF17
Maç Skoru_0:3	FF18
Maç Skoru_1:3	FF19
Maç Skoru_2:3	FF20
Maç Skoru_1:4	FF21
Maç Skoru_2:4	FF22
Maç Skoru_diğer	FF23
Maç Skoru_0:5	FF24
Maç Skoru_5:0	FF25
Maç Skoru_0:6	FF26
Maç Skoru_6:0	FF27
"""


def temiz_text(s):
    if s is None:
        return ""

    s = str(s).strip()

    ceviri = {
        "ı": "i",
        "İ": "i",
        "ş": "s",
        "Ş": "s",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }

    for a, b in ceviri.items():
        s = s.replace(a, b)

    s = s.lower()
    s = s.replace(",", ".")
    s = s.replace("üstü", "üst")
    s = s.replace("ustu", "ust")
    s = s.replace(" / ", "/")
    s = s.replace(" - ", "-")
    s = re.sub(r"\s+", " ", s)

    return s.strip()


def norm_key(s):
    s = temiz_text(s)
    s = s.replace(" ve ", " ")
    s = re.sub(r"[^a-z0-9./:+_\- ]+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def olasi_anahtarlar(veri_adi):
    """
    İddaa JSON bazen beraberlik için X,
    senin şablonda ise 0 kullanıyor olabilir.
    Bu yüzden alternatif anahtarlar üretir.
    """
    adaylar = set()

    v = veri_adi.strip()
    adaylar.add(v)

    # 0 / X dönüşümleri
    adaylar.add(v.replace("_0", "_X"))
    adaylar.add(v.replace("_X", "_0"))
    adaylar.add(v.replace("/0", "/X"))
    adaylar.add(v.replace("/X", "/0"))
    adaylar.add(v.replace(" 0 ", " X "))
    adaylar.add(v.replace(" X ", " 0 "))

    # Alt/Üstü - Alt/Üst farkı
    adaylar.add(v.replace("Alt/Üstü", "Alt/Üst"))
    adaylar.add(v.replace("Alt/Üst", "Alt/Üstü"))

    # Üst / Ust
    adaylar.add(v.replace("Üst", "Ust"))
    adaylar.add(v.replace("Ust", "Üst"))

    return [norm_key(x) for x in adaylar]


def sablon_oku():
    """
    Önce oran_sablonu.txt varsa onu okur.
    Yoksa yukarıdaki VARSAYILAN_SABLON listesini kullanır.
    """
    if os.path.exists(SABLON_TXT):
        print(f"📄 Şablon okunuyor: {SABLON_TXT}")
        with open(SABLON_TXT, "r", encoding="utf-8") as f:
            text = f.read()
    else:
        print("⚠️ oran_sablonu.txt bulunamadı, kod içindeki varsayılan liste kullanılacak.")
        text = VARSAYILAN_SABLON

    kolonlar = []

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        if line.lower().startswith("veri adı"):
            continue

        # Tab, noktalı virgül veya çoklu boşluğa göre ayır
        if "\t" in line:
            parts = line.rsplit("\t", 1)
        elif ";" in line:
            parts = line.rsplit(";", 1)
        else:
            parts = re.split(r"\s{2,}", line)
            if len(parts) < 2:
                continue
            parts = [parts[0], parts[-1]]

        if len(parts) != 2:
            continue

        veri_adi = parts[0].strip()
        kod = parts[1].strip()

        if veri_adi and kod:
            kolonlar.append((veri_adi, kod))

    return kolonlar


def json_oku():
    if not os.path.exists(JSON_DOSYA):
        raise FileNotFoundError(f"JSON bulunamadı: {JSON_DOSYA}")

    with open(JSON_DOSYA, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        return data.get("matches", [])

    if isinstance(data, list):
        return data

    return []


def oran_index_hazirla(oranlar):
    """
    JSON içindeki oranlar için normalize edilmiş hızlı arama sözlüğü oluşturur.
    """
    idx = {}

    for k, v in (oranlar or {}).items():
        nk = norm_key(k)
        idx[nk] = v

        # X / 0 alternatifleri
        idx[nk.replace("_x", "_0")] = v
        idx[nk.replace("_0", "_x")] = v
        idx[nk.replace("/x", "/0")] = v
        idx[nk.replace("/0", "/x")] = v

    return idx


def oran_bul(oran_idx, veri_adi):
    for aday in olasi_anahtarlar(veri_adi):
        if aday in oran_idx:
            return oran_idx[aday]

    return ""


def excel_yaz(maclar, kolonlar):
    wb = Workbook()
    ws = wb.active
    ws.title = "Maclar"

    bilgi_kolonlari = [
        ("Tarih", "tarih"),
        ("Saat", "saat"),
        ("Lig", "lig"),
        ("Ev Sahibi", "ev_sahibi"),
        ("Deplasman", "deplasman"),
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    col = 1

    # Temel maç bilgileri
    for baslik, _key in bilgi_kolonlari:
        c1 = ws.cell(1, col, baslik)
        c2 = ws.cell(2, col, "")

        c1.fill = header_fill
        c1.font = white_font
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = thin

        c2.fill = sub_fill
        c2.border = thin

        col += 1

    # Oran kolonları
    for veri_adi, kod in kolonlar:
        c1 = ws.cell(1, col, kod)
        c2 = ws.cell(2, col, veri_adi)

        c1.fill = header_fill
        c1.font = white_font
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = thin

        c2.fill = sub_fill
        c2.font = bold_font
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2.border = thin

        col += 1

    # Maç satırları
    row_no = 3

    for m in maclar:
        oran_idx = oran_index_hazirla(m.get("oranlar", {}))

        col = 1

        for _baslik, key in bilgi_kolonlari:
            ws.cell(row_no, col, m.get(key, ""))
            ws.cell(row_no, col).border = thin
            col += 1

        for veri_adi, _kod in kolonlar:
            val = oran_bul(oran_idx, veri_adi)

            try:
                if val != "":
                    val = float(val)
            except Exception:
                pass

            cell = ws.cell(row_no, col, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if isinstance(val, float):
                cell.number_format = "0.00"

            col += 1

        row_no += 1

    # Genişlikler
    widths = {
        1: 12,
        2: 8,
        3: 28,
        4: 24,
        5: 24,
    }

    total_cols = len(bilgi_kolonlari) + len(kolonlar)

    for i in range(1, total_cols + 1):
        letter = get_column_letter(i)

        if i in widths:
            ws.column_dimensions[letter].width = widths[i]
        else:
            ws.column_dimensions[letter].width = 10

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 45

    ws.freeze_panes = "F3"
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(EXCEL_DOSYA), exist_ok=True)

    try:
        wb.save(EXCEL_DOSYA)
        print(f"✅ Excel kaydedildi: {EXCEL_DOSYA}")
    except PermissionError:
        yeni = EXCEL_DOSYA.replace(".xlsx", "_yeni.xlsx")
        wb.save(yeni)
        print("⚠️ Eski Excel dosyası açık olduğu için yeni dosya yazıldı:")
        print(yeni)


def main():
    print("📖 JSON okunuyor...")
    maclar = json_oku()

    if not maclar:
        print("❌ JSON içinde maç bulunamadı.")
        return

    print(f"✅ {len(maclar)} maç bulundu.")

    kolonlar = sablon_oku()

    if not kolonlar:
        print("❌ Şablon boş. oran_sablonu.txt içine kolon listesini ekle.")
        return

    print(f"✅ {len(kolonlar)} oran kolonu kullanılacak.")
    print("📊 Excel oluşturuluyor...")

    excel_yaz(maclar, kolonlar)

    try:
        os.startfile(EXCEL_DOSYA)
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("❌ HATA:", e)

    input("Çıkmak için Enter...")